#!/usr/bin/env python3
"""One-time fix: correct orgmap (ministry / MOA / clearing) from the workbook's
**Source** sheet, columns H:K, IN PLACE — no database rebuild, no data loss.

Why this exists: the Source sheet has TWO ministry columns (J = org->ministry, the
correct one; M = ministry->target). The original importer's header lookup let column M
overwrite column J, so orgmap.ministry was wrong/blank for most codes. This reads column
J directly and corrects orgmap, then aligns existing entries to match.

A safety backup is written to server/backups/ before any change.

Usage:
    python3 fix_orgmap.py "2026 Collection Database (Final_).xlsx"
    python3 fix_orgmap.py "<workbook>.xlsx" --no-regroup   # correct orgmap only
"""
import argparse
import datetime
import getpass
import pathlib
import sqlite3
import sys

import openpyxl

HERE = pathlib.Path(__file__).resolve().parent
DB = HERE / "collection.db"


def load_source(path):
    """{org_code: (clearing, ministry, moa)} from the Source sheet's H:K columns.
    Validates the header first, so a shifted/edited sheet aborts loudly instead of
    silently rewriting orgmap with data from the wrong columns."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Source" not in wb.sheetnames:
        sys.exit("Workbook has no 'Source' sheet.")
    rows = list(wb["Source"].iter_rows(values_only=True))
    wb.close()

    def cell(row, i):
        return str(row[i]).strip().upper() if len(row) > i and row[i] is not None else ""

    hdr = None
    for ri, row in enumerate(rows):
        if cell(row, 7) == "ORGANIZATIONAL CODE":
            if cell(row, 9) != "MINISTRY" or cell(row, 10) != "MOA":
                sys.exit("Source sheet columns look shifted (expected H=ORGANIZATIONAL CODE, "
                         "J=MINISTRY, K=MOA). Aborting — no changes made to the database.")
            hdr = ri
            break
    if hdr is None:
        sys.exit("Could not find an 'ORGANIZATIONAL CODE' header in column H of the Source sheet.")

    out = {}
    for row in rows[hdr + 1:]:
        if len(row) > 10 and row[7]:
            code = str(row[7]).strip()
            if code:
                out[code] = (str(row[8] or "").strip(), str(row[9] or "").strip(), str(row[10] or "").strip())
    return out


def backup(ts):
    (HERE / "backups").mkdir(exist_ok=True)
    dest = HERE / "backups" / f"pre-fix-orgmap-{ts}.db"
    if dest.exists():
        sys.exit(f"Backup {dest.name} already exists — refusing to overwrite it. "
                 "Pass a different --ts, or move the old backup aside.")
    src = sqlite3.connect(str(DB))
    dst = sqlite3.connect(str(dest))
    with dst:
        src.backup(dst)
    dst.close()
    src.close()
    return dest


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="path to the Collection Database workbook")
    ap.add_argument("--no-regroup", action="store_true",
                    help="only correct orgmap; do NOT touch existing entries")
    ap.add_argument("--ts", default=datetime.datetime.now().strftime("%Y%m%d-%H%M%S"),
                    help="label for the backup filename (default: current date-time)")
    a = ap.parse_args()
    if not DB.exists():
        sys.exit("No collection.db found next to this script.")

    src = load_source(a.xlsx)
    print(f"Read {len(src)} org codes from Source!H:K.")
    b = backup(a.ts)
    print(f"Safety backup written: {b.name}")

    con = sqlite3.connect(str(DB))
    con.row_factory = sqlite3.Row
    # map trimmed org_code -> real stored org_code (handles stray trailing spaces)
    stored = {r["org_code"].strip(): r["org_code"] for r in con.execute("SELECT org_code FROM orgmap")}
    corrected = added = regrouped = 0
    for code, (clearing, ministry, moa) in src.items():
        real = stored.get(code)
        if real is None:
            con.execute("INSERT INTO orgmap(org_code,clearing,ministry,moa) VALUES(?,?,?,?)",
                        (code, clearing, ministry, moa))
            added += 1
        else:
            r = con.execute("SELECT clearing,ministry,moa FROM orgmap WHERE org_code=?", [real]).fetchone()
            if (r["clearing"] or "") != clearing or (r["ministry"] or "") != ministry or (r["moa"] or "") != moa:
                con.execute("UPDATE orgmap SET clearing=?,ministry=?,moa=? WHERE org_code=?",
                            (clearing, ministry, moa, real))
                corrected += 1
        if not a.no_regroup and ministry:
            cur = con.execute("UPDATE transactions SET ministry=? "
                              "WHERE trim(org_code)=? AND ministry<>? AND tagging!='Target'",
                              (ministry, code, ministry))
            regrouped += cur.rowcount
    by_user = "console:" + getpass.getuser()
    con.execute("INSERT INTO audit(action,detail,by_user) VALUES('Fix Orgmap',?,?)",
                [f"Source!H:K import: {corrected} corrected, {added} added, {regrouped} entries regrouped", by_user])
    con.commit()
    total = con.execute("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE tagging!='Target'").fetchone()[0]
    blanks = con.execute("SELECT COUNT(*) FROM orgmap WHERE ministry IS NULL OR ministry=''").fetchone()[0]
    con.close()
    print(f"orgmap: {corrected} corrected, {added} added.  Entries re-grouped: {regrouped}.")
    print(f"Org codes still without a ministry: {blanks}")
    print(f"Grand total (unchanged): PHP {total:,.2f}")
    print("Done. Restart the app to see it, or it applies on the next page load.")


if __name__ == "__main__":
    main()
