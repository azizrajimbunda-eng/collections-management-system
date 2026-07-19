#!/usr/bin/env python3
"""Import the office's Excel workbook into collection.db.

Reads the **Database** sheet (transactions) and the **Source** sheet (org-code map,
lookup lists, ministry targets), converts them to the internal JSON shape, then runs
import_data.py so the certification config / roles / mapping are applied exactly as usual.

Usage:
    python3 import_xlsx.py "/path/to/2026 Collection Database (Final_).xlsx"
    python3 import_xlsx.py "<file>" --admin-pass "myNewPassword"
"""
import argparse
import datetime
import json
import pathlib
import subprocess
import sys

import openpyxl

HERE = pathlib.Path(__file__).resolve().parent


def cval(v):
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    return v


def find_header(sheet, marker="ORGANIZATIONAL CODE"):
    """Return (iterator positioned after the header, {NAME: col_index})."""
    it = sheet.iter_rows(values_only=True)
    for row in it:
        if row and any(str(c).strip().upper() == marker for c in row if c):
            idx = {}
            for j, c in enumerate(row):
                if c not in (None, ""):
                    # keep the FIRST column for a given header name — the Source sheet has two
                    # "MINISTRY" columns (J = org->ministry, M = ministry->target); J must win.
                    idx.setdefault(str(c).strip().upper(), j)
            return it, idx
    sys.exit(f"Could not find a header row containing '{marker}' in sheet '{sheet.title}'.")


def load_workbook_data(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Database" not in wb.sheetnames or "Source" not in wb.sheetnames:
        sys.exit("Workbook must contain 'Database' and 'Source' sheets.")

    # ----- Database sheet -> transactions -----
    it, H = find_header(wb["Database"])

    def g(row, name):
        i = H.get(name)
        return cval(row[i]) if (i is not None and len(row) > i) else ""

    txns = []
    for row in it:
        org = g(row, "ORGANIZATIONAL CODE")
        if not org:
            continue
        txns.append({
            "org_code": org, "clearing": g(row, "CLEARING ACCOUNT"),
            "ministry": g(row, "MINISTRY"), "office": g(row, "OFFICE"),
            "payment_method": g(row, "PAYMENT METHOD"), "lbp_branch": g(row, "BANK BRANCH"),
            "amount": g(row, "AMOUNT"), "date": g(row, "TRANSACTION DATE"),
            "type": g(row, "TYPE OF COLLECTION"), "tagging": g(row, "TAGGING") or "Actual",
            "remarks": g(row, "REMARKS"),
        })

    # ----- Source sheet -> orgmap, lists, targets -----
    it2, S = find_header(wb["Source"])

    def col(name):
        return S.get(name)

    def sv(row, ci):
        return cval(row[ci]) if (ci is not None and len(row) > ci) else ""

    offices, pays, branches, months, types, orgmap, targets = [], [], [], [], [], [], []
    c_off, c_pay, c_br, c_mon, c_typ = (col("OFFICE"), col("MODE OF PAYMENT"),
                                        col("BANK BRANCH"), col("MONTH"), col("TYPE OF COLLECTION"))
    c_org, c_clr, c_min, c_moa = (col("ORGANIZATIONAL CODE"), col("CLEARING ACCOUNT"),
                                  col("MINISTRY"), col("MOA"))
    c_tamt = col("2026 TARGET AMOUNT")
    c_tmin = (c_tamt - 1) if c_tamt is not None else None   # ministry sits just left of target
    for row in it2:
        if sv(row, c_off):
            offices.append(sv(row, c_off))
        if sv(row, c_pay):
            pays.append(sv(row, c_pay))
        if sv(row, c_br):
            branches.append(sv(row, c_br))
        if sv(row, c_mon):
            months.append(sv(row, c_mon))
        if sv(row, c_typ):
            types.append(sv(row, c_typ))
        oc = sv(row, c_org)
        if oc:
            orgmap.append({"org_code": oc, "clearing": sv(row, c_clr),
                           "ministry": sv(row, c_min), "moa": sv(row, c_moa)})
        tm = sv(row, c_tmin)
        if tm and sv(row, c_tamt) != "":
            try:
                targets.append({"ministry": tm, "target": float(row[c_tamt] or 0)})
            except (TypeError, ValueError):
                pass

    seed = {
        "transactions": txns, "orgmap": orgmap, "targets": targets,
        "lists": {
            "offices": offices, "payment_methods": pays, "lbp_branches": branches,
            "months": months, "collection_types": types,
            "org_codes": [o["org_code"] for o in orgmap],
            "ministries": sorted({o["ministry"] for o in orgmap if o["ministry"]}),
        },
    }
    return seed


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="path to the Excel workbook")
    ap.add_argument("--admin-pass", default=None,
                    help="admin password; if omitted a strong random one is generated and printed")
    a = ap.parse_args()

    seed = load_workbook_data(a.xlsx)
    tmp = HERE / "imported_from_xlsx.json"
    tmp.write_text(json.dumps(seed), encoding="utf-8")
    print(f"Converted {len(seed['transactions'])} transactions from the workbook.")
    print("Building the database (applying certification config, roles, mapping)...")
    cmd = [sys.executable, str(HERE / "import_data.py"), "--seed", str(tmp), "--force"]
    if a.admin_pass:
        cmd += ["--admin-pass", a.admin_pass]
    r = subprocess.run(cmd, cwd=str(HERE))
    sys.exit(r.returncode)


if __name__ == "__main__":
    main()
