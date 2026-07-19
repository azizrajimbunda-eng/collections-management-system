#!/usr/bin/env python3
"""Export collection.db back to an Excel workbook (Database + Source sheets).

Produces a file that the office can open in Excel, and that import_xlsx.py can re-import.

Usage:
    python3 export_xlsx.py                 # -> exports/collection-export-<timestamp>.xlsx
    python3 export_xlsx.py output.xlsx
"""
import datetime
import pathlib
import sqlite3
import sys

import openpyxl

HERE = pathlib.Path(__file__).resolve().parent
DB = HERE / "collection.db"

DB_HEADERS = ["ORGANIZATIONAL CODE", "CLEARING ACCOUNT", "MINISTRY", "OFFICE", "PAYMENT METHOD",
              "BANK BRANCH", "AMOUNT", "YEAR", "MONTH", "DAY", "Transaction Date",
              "TYPE OF COLLECTION", "TAGGING", "REMARKS"]


# Characters that make Excel/LibreOffice treat a cell as a live formula when it
# appears first. Prefixing with an apostrophe forces the cell to be read as text.
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _safe(v):
    """Neutralize spreadsheet formula/CSV injection in a user-supplied string.

    Non-strings (numbers, dates, None) pass through unchanged.
    """
    if isinstance(v, str) and v and v[0] in _FORMULA_PREFIXES:
        return "'" + v
    return v


def build_workbook(con):
    """Build the Database + Source workbook from an open DB connection. Reused by the app."""
    con.row_factory = sqlite3.Row
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Database"
    ws.append(DB_HEADERS)
    for t in con.execute("SELECT * FROM transactions ORDER BY txn_date, id"):
        ws.append([_safe(t["org_code"]), _safe(t["clearing"]), _safe(t["ministry"]),
                   _safe(t["office"]), _safe(t["payment_method"]), _safe(t["lbp_branch"]),
                   t["amount"], t["year"], t["month"], t["day"], t["txn_date"],
                   _safe(t["type"]), _safe(t["tagging"]), _safe(t["remarks"])])

    src = wb.create_sheet("Source")
    lists = {ln: [r[0] for r in con.execute(
        "SELECT value FROM list_items WHERE list_name=? ORDER BY ord", [ln])]
        for ln in ["offices", "payment_methods", "lbp_branches", "months", "collection_types"]}
    orgmap = list(con.execute("SELECT org_code, clearing, ministry, moa FROM orgmap ORDER BY org_code"))
    targets = list(con.execute("SELECT ministry, target FROM targets ORDER BY ministry"))
    src.append(["Office", "Mode of Payment", "Bank Branch", "Month", "Type of Collection", "",
                "ORGANIZATIONAL CODE", "CLEARING ACCOUNT", "MINISTRY", "MOA", "",
                "Ministry", "2026 Target Amount"])
    keys = ["offices", "payment_methods", "lbp_branches", "months", "collection_types"]
    n = max([len(v) for v in lists.values()] + [len(orgmap), len(targets)] + [0])
    for i in range(n):
        row = [""] * 13
        for ci, k in enumerate(keys):
            if i < len(lists[k]):
                row[ci] = _safe(lists[k][i])
        if i < len(orgmap):
            o = orgmap[i]
            row[6], row[7], row[8], row[9] = (_safe(o["org_code"]), _safe(o["clearing"]),
                                              _safe(o["ministry"]), _safe(o["moa"]))
        if i < len(targets):
            tg = targets[i]
            row[11], row[12] = _safe(tg["ministry"]), tg["target"]
        src.append(row)
    return wb


def build_official_report_workbook(d, year):
    """Build a formatted Excel version of the "Collection Report (Official Format)"
    (db.report_official's payload), preserving the on-screen nested I/II/III/IV structure:
    indentation per depth, bold+shaded section/subtotal rows, matching officialHTML() in
    static/index.html."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Collection Report"

    cols = d["cols"]
    ncols = 1 + len(cols) + 1  # Particulars + months + TOTAL
    bold = openpyxl.styles.Font(bold=True)
    center = openpyxl.styles.Alignment(horizontal="center")
    fill_grp = openpyxl.styles.PatternFill("solid", fgColor="F1F4F9")
    fill_tot = openpyxl.styles.PatternFill("solid", fgColor="DFE7F3")

    for row, text, size in ((1, "COSTA VERDE AUTONOMOUS REGION", 12),
                             (2, "SUMMARY OF COLLECTION", 12), (3, f"FY {year or 'All Years'}", 11)):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        cell = ws.cell(row, 1, text)
        cell.font = openpyxl.styles.Font(bold=True, size=size)
        cell.alignment = center

    header_row = 5
    ws.cell(header_row, 1, "Particulars").font = bold
    for i, c in enumerate(cols):
        ws.cell(header_row, 2 + i, c["label"]).font = bold
    ws.cell(header_row, ncols, "TOTAL").font = bold
    for c in range(1, ncols + 1):
        cell = ws.cell(header_row, c)
        cell.fill = fill_grp
        cell.alignment = center

    r = header_row + 1

    def write_node(n, depth):
        nonlocal r
        has_kids = bool(n.get("children"))
        emphasize = n.get("isTotal") or has_kids
        label_cell = ws.cell(r, 1, _safe(n["label"]))
        label_cell.alignment = openpyxl.styles.Alignment(indent=depth, wrap_text=True)
        if emphasize:
            label_cell.font = bold
        for i, c in enumerate(cols):
            v = ws.cell(r, 2 + i, n["cells"].get(c["key"], 0))
            v.number_format = "#,##0.00"
            if emphasize:
                v.font = bold
        t = ws.cell(r, ncols, n["total"])
        t.number_format = "#,##0.00"
        t.font = bold
        if n.get("isTotal"):
            for c in range(1, ncols + 1):
                ws.cell(r, c).fill = fill_tot
        elif has_kids:
            for c in range(1, ncols + 1):
                ws.cell(r, c).fill = fill_grp
        r += 1
        for child in n.get("children") or []:
            write_node(child, depth + 1)

    for sec in d["sections"]:
        write_node(sec, 0)

    ws.freeze_panes = ws.cell(header_row + 1, 1).coordinate
    ws.column_dimensions["A"].width = 46
    for i in range(len(cols)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(2 + i)].width = 13
    ws.column_dimensions[openpyxl.utils.get_column_letter(ncols)].width = 16
    return wb


def main():
    if not DB.exists():
        sys.exit("No collection.db to export.")
    con = sqlite3.connect(str(DB))
    wb = build_workbook(con)
    txn_rows = wb["Database"].max_row - 1
    con.close()
    out = (sys.argv[1] if len(sys.argv) > 1 else
           str(HERE / "exports" / ("collection-export-" +
               datetime.datetime.now().strftime("%Y%m%d-%H%M%S") + ".xlsx")))
    pathlib.Path(out).parent.mkdir(exist_ok=True)
    wb.save(out)
    print(f"Exported to: {out}\n  transactions: {txn_rows}")


if __name__ == "__main__":
    main()
